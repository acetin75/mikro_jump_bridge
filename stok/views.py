import logging
from decimal import Decimal

from django.contrib import messages
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.utils import timezone
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from .degerleme import urun_degerleme
from .forms import StokHareketiForm, UrunForm
from .models import DEGERLEME_YONTEMLERI, StokHareketi, Urun

logger = logging.getLogger("muhasebe")


# ── Ürün Listesi ──────────────────────────────────────────────────────────
def urun_listesi(request):
    q = request.GET.get("q", "").strip()
    sadece_aktif = request.GET.get("aktif", "") != "0"
    kritik = request.GET.get("kritik", "")

    qs = Urun.objects.all()
    if sadece_aktif:
        qs = qs.filter(aktif=True)
    if q:
        qs = qs.filter(ad__icontains=q) | qs.filter(kod__icontains=q)

    urunler = list(qs.order_by("ad"))
    if kritik:
        urunler = [u for u in urunler if u.kritik_stok_mu]

    return render(request, "stok/list.html", {
        "urunler": urunler,
        "q": q,
        "sadece_aktif": sadece_aktif,
        "kritik": kritik,
    })


# ── Ürün Ekle ─────────────────────────────────────────────────────────────
def urun_ekle(request):
    form = UrunForm(request.POST or None)
    if request.method == "POST" and form.is_valid():
        urun = form.save()
        logger.info("Ürün oluşturuldu: %s (pk=%s)", urun.ad, urun.pk)
        messages.success(request, f"Ürün '{urun.ad}' oluşturuldu.")
        return redirect("stok:urun_detay", pk=urun.pk)
    return render(request, "stok/form.html", {"form": form, "baslik": "Yeni Ürün"})


# ── Ürün Detay ────────────────────────────────────────────────────────────
def urun_detay(request, pk):
    urun = get_object_or_404(Urun, pk=pk)
    hareketler = urun.hareketler.all()[:50]
    hareket_form = StokHareketiForm(initial={"urun": urun, "tarih": timezone.now().date()})
    degerleme = urun_degerleme(urun)
    return render(request, "stok/detail.html", {
        "urun": urun,
        "hareketler": hareketler,
        "hareket_form": hareket_form,
        "degerleme": degerleme,
    })


# ── Ürün Düzenle ──────────────────────────────────────────────────────────
def urun_duzenle(request, pk):
    urun = get_object_or_404(Urun, pk=pk)
    form = UrunForm(request.POST or None, instance=urun)
    if request.method == "POST" and form.is_valid():
        form.save()
        messages.success(request, "Ürün güncellendi.")
        return redirect("stok:urun_detay", pk=urun.pk)
    return render(request, "stok/form.html", {"form": form, "baslik": f"Düzenle: {urun.ad}", "urun": urun})


# ── Ürün Sil ──────────────────────────────────────────────────────────────
def urun_sil(request, pk):
    urun = get_object_or_404(Urun, pk=pk)
    if request.method == "POST":
        try:
            ad = urun.ad
            urun.delete()
            logger.info("Ürün silindi: %s", ad)
            messages.success(request, f"Ürün '{ad}' silindi.")
            return redirect("stok:list")
        except Exception as e:
            logger.error("Ürün silme hatası: %s", e, exc_info=True)
            messages.error(request, "Ürün silinemedi — stok hareketi veya fatura kalemi bağlı olabilir.")
            return redirect("stok:urun_detay", pk=pk)
    return render(request, "confirm_delete.html", {
        "nesne": urun,
        "geri_url": "stok:list",
    })


# ── Stok Hareketi Ekle ────────────────────────────────────────────────────
def stok_hareketi_ekle(request, urun_pk=None):
    urun = get_object_or_404(Urun, pk=urun_pk) if urun_pk else None
    initial = {"tarih": timezone.now().date()}
    if urun:
        initial["urun"] = urun
    form = StokHareketiForm(request.POST or None, initial=initial)
    if request.method == "POST" and form.is_valid():
        hareket = form.save()
        logger.info("Stok hareketi: %s %s %s", hareket.get_tip_display(), hareket.urun, hareket.miktar)
        messages.success(request, "Stok hareketi kaydedildi.")
        return redirect("stok:urun_detay", pk=hareket.urun_id)
    return render(request, "stok/hareket_form.html", {
        "form": form,
        "baslik": "Stok Hareketi Ekle",
        "urun": urun,
    })


# ── Stok Hareketi Sil ─────────────────────────────────────────────────────
def stok_hareketi_sil(request, pk):
    hareket = get_object_or_404(StokHareketi, pk=pk)
    urun_pk = hareket.urun_id
    if request.method == "POST":
        hareket.delete()
        messages.success(request, "Hareket silindi.")
        return redirect("stok:urun_detay", pk=urun_pk)
    return render(request, "confirm_delete.html", {
        "nesne": hareket,
        "geri_url": "stok:list",
    })


# ── Stok Değerleme Raporu ───────────────────────────────────────
def stok_degerleme_raporu(request):
    """Üm aktüm ürünler için seçili yöntemle maliyet raporu."""
    yontem = request.GET.get("yontem", "")

    qs = Urun.objects.filter(aktif=True).prefetch_related("hareketler")
    if yontem:
        qs = qs.filter(degerleme_yontemi=yontem)

    rapor = []
    toplam_deger = Decimal("0")
    toplam_alis_degeri = Decimal("0")

    for urun in qs:
        bilgi = urun_degerleme(urun)
        kar_marji = None
        kar_tutari = None
        if bilgi["birim_maliyet"] > 0 and urun.satis_fiyati > 0:
            kar_tutari = urun.satis_fiyati - bilgi["birim_maliyet"]
            kar_marji = float(kar_tutari / bilgi["birim_maliyet"] * 100)
        rapor.append({
            "urun": urun,
            **bilgi,
            "kar_marji": kar_marji,
            "kar_tutari": kar_tutari,
        })
        toplam_deger += bilgi["toplam_deger"]
        if bilgi["mevcut_adet"] > 0 and bilgi["birim_maliyet"] > 0:
            toplam_alis_degeri += bilgi["toplam_deger"]

    # Portföy satış değeri
    toplam_satis_degeri = sum(
        float(r["urun"].satis_fiyati) * float(r["mevcut_adet"])
        for r in rapor
        if r["mevcut_adet"] > 0
    )

    return render(request, "stok/degerleme.html", {
        "rapor": rapor,
        "toplam_deger": toplam_deger,
        "toplam_satis_degeri": toplam_satis_degeri,
        "yontem": yontem,
        "yontem_secenekler": DEGERLEME_YONTEMLERI,
    })


# ── Stok Değerleme Excel ───────────────────────────────────────
def stok_degerleme_excel(request):
    """Stok değerleme raporunu Excel'e aktar."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Stok Değerleme"

    tarih_str = timezone.now().strftime("%d.%m.%Y %H:%M")
    ws["A1"] = f"Stok Değerleme Raporu — {tarih_str}"
    ws["A1"].font = Font(bold=True, size=13)
    ws.merge_cells("A1:J1")

    basliklar = [
        "Kod", "Ürün Adı", "Birim", "Mevcut Stok",
        "Yöntem", "Birim Maliyet (₺)", "Toplam Stok Değeri (₺)",
        "Satış Fiyatı (₺)", "Satış Değeri (₺)", "Kâr Marjı (%)",
    ]
    fill_h = PatternFill(fill_type="solid", fgColor="1F2D3D")
    font_h = Font(color="FFFFFF", bold=True)
    for i, h in enumerate(basliklar, 1):
        c = ws.cell(3, i, h)
        c.fill = fill_h
        c.font = font_h
        c.alignment = Alignment(horizontal="center")

    urunler = Urun.objects.filter(aktif=True).prefetch_related("hareketler")
    toplam_maliyet = Decimal("0")
    toplam_satis = Decimal("0")

    for urun in urunler:
        bilgi = urun_degerleme(urun)
        satis_degeri = urun.satis_fiyati * bilgi["mevcut_adet"]
        kar_marji = None
        if bilgi["birim_maliyet"] > 0 and urun.satis_fiyati > 0:
            kar_marji = round(
                float((urun.satis_fiyati - bilgi["birim_maliyet"]) / bilgi["birim_maliyet"] * 100), 2
            )
        toplam_maliyet += bilgi["toplam_deger"]
        toplam_satis += satis_degeri
        ws.append([
            urun.kod or "",
            urun.ad,
            urun.birim,
            float(bilgi["mevcut_adet"]),
            bilgi["yontem_adi"],
            float(bilgi["birim_maliyet"]),
            float(bilgi["toplam_deger"]),
            float(urun.satis_fiyati),
            float(satis_degeri),
            kar_marji,
        ])

    # Toplam satırı
    fill_t = PatternFill(fill_type="solid", fgColor="FFF3CD")
    t_row = ws.max_row + 1
    ws.cell(t_row, 1, "TOPLAM").font = Font(bold=True)
    ws.cell(t_row, 7, float(toplam_maliyet)).font = Font(bold=True)
    ws.cell(t_row, 9, float(toplam_satis)).font = Font(bold=True)
    for c in ws[t_row]:
        c.fill = fill_t

    # FIFO Katman Detayı — 2. sayfa
    ws2 = wb.create_sheet("FIFO Katmanlar")
    ws2["A1"] = "FIFO Stok Katmanları (Değerleme Yöntemi: FIFO olan ürünler)"
    ws2["A1"].font = Font(bold=True, size=12)
    ws2.merge_cells("A1:F1")
    for i, h in enumerate(["Köd", "Ürün", "Giriş Tarihi", "Miktar", "Birim Fiyat (₺)", "Katman Değeri (₺)"], 1):
        c = ws2.cell(3, i, h)
        c.fill = fill_h
        c.font = font_h
        c.alignment = Alignment(horizontal="center")

    fifo_urunler = Urun.objects.filter(aktif=True, degerleme_yontemi="fifo").prefetch_related("hareketler")
    for urun in fifo_urunler:
        bilgi = urun_degerleme(urun)
        for katman in bilgi.get("fifo_katmanlar", []):
            ws2.append([
                urun.kod or "",
                urun.ad,
                str(katman["tarih"]),
                float(katman["miktar"]),
                float(katman["birim_fiyat"]),
                float(katman["deger"]),
            ])

    for sheet in [ws, ws2]:
        for col in sheet.columns:
            max_len = max((len(str(c.value or "")) for c in col), default=0)
            sheet.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 50)

    resp = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    fname = f"stok_degerleme_{timezone.now().strftime('%Y%m%d')}.xlsx"
    resp["Content-Disposition"] = f"attachment; filename={fname}"
    wb.save(resp)
    return resp
