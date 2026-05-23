from django.shortcuts import render
from django.http import HttpResponse
from django.utils import timezone
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from decimal import Decimal

from cari.models import Cari, HesapHareketi
from tahsilat.models import Tahsilat
from gider.models import Gider
from fatura.models import Fatura, FaturaKalemi


# ── yardımcı: sütunları otomatik genişlet ─────────────────────────────────
def _auto_width(ws):
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=0)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 50)


# ── yardımcı: başlık satırı ───────────────────────────────────────────────
def _header_row(ws, cols, row=1):
    fill = PatternFill(fill_type="solid", fgColor="1F2D3D")
    font = Font(color="FFFFFF", bold=True)
    for i, col in enumerate(cols, 1):
        c = ws.cell(row=row, column=i, value=col)
        c.fill = fill
        c.font = font
        c.alignment = Alignment(horizontal="center")


# ── Cari Bakiye Listesi ───────────────────────────────────────────────────
def excel_cari_bakiye(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Cari Bakiyeler"

    baslik_cols = ["#", "Ünvan", "Tip", "Telefon", "Email", "Borç", "Alacak", "Bakiye"]
    _header_row(ws, baslik_cols)

    cariler = Cari.objects.all().order_by("unvan")
    for i, cari in enumerate(cariler, 2):
        borclu = sum(h.borc for h in cari.hareketler.all()) or Decimal("0")
        alacakli = sum(h.alacak for h in cari.hareketler.all()) or Decimal("0")
        bakiye = alacakli - borclu
        ws.append([
            "", cari.ad, cari.get_tip_display(),
            cari.telefon or "", cari.email or "",
            float(borclu), float(alacakli), float(bakiye),
        ])

    _auto_width(ws)
    resp = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    resp["Content-Disposition"] = "attachment; filename=cari_bakiye.xlsx"
    wb.save(resp)
    return resp


# ── Cari Ekstresi ─────────────────────────────────────────────────────────
def excel_cari_ekstre(request):
    cari_id = request.GET.get("cari_id")
    bas = request.GET.get("bas")
    bitis = request.GET.get("bitis")

    qs = HesapHareketi.objects.select_related("cari").all().order_by("tarih")
    if cari_id:
        qs = qs.filter(cari_id=cari_id)
    if bas:
        qs = qs.filter(tarih__gte=bas)
    if bitis:
        qs = qs.filter(tarih__lte=bitis)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Cari Ekstresi"
    _header_row(ws, ["Tarih", "Cari", "Açıklama", "Tür", "Tutar"])

    for h in qs:
        ws.append([
            str(h.tarih), h.cari.ad, h.aciklama,
            h.get_hareket_tipi_display(), float(h.tutar),
        ])

    _auto_width(ws)
    resp = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    resp["Content-Disposition"] = "attachment; filename=cari_ekstre.xlsx"
    wb.save(resp)
    return resp


# ── Tahsilat Raporu ───────────────────────────────────────────────────────
def excel_tahsilat(request):
    bas = request.GET.get("bas")
    bitis = request.GET.get("bitis")

    qs = Tahsilat.objects.select_related("cari").order_by("-tarih")
    if bas:
        qs = qs.filter(tarih__gte=bas)
    if bitis:
        qs = qs.filter(tarih__lte=bitis)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Tahsilatlar"
    _header_row(ws, ["Tarih", "Cari", "Açıklama", "Tutar", "Ödeme Yöntemi"])

    for t in qs:
        ws.append([
            str(t.tarih), t.cari.ad, t.aciklama or "",
            float(t.tutar), getattr(t, "odeme_yontemi", ""),
        ])

    _auto_width(ws)
    resp = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    resp["Content-Disposition"] = "attachment; filename=tahsilat_raporu.xlsx"
    wb.save(resp)
    return resp


# ── Gider Raporu ──────────────────────────────────────────────────────────
def excel_gider(request):
    bas = request.GET.get("bas")
    bitis = request.GET.get("bitis")

    qs = Gider.objects.select_related("kategori").order_by("-tarih")
    if bas:
        qs = qs.filter(tarih__gte=bas)
    if bitis:
        qs = qs.filter(tarih__lte=bitis)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Giderler"
    _header_row(ws, ["Tarih", "Kategori", "Açıklama", "KDV Hariç", "KDV %", "KDV Dahil", "Belge No"])

    for g in qs:
        ws.append([
            str(g.tarih),
            str(g.kategori) if g.kategori else "",
            g.aciklama,
            float(g.kdv_haric_tutar),
            g.kdv_orani,
            float(g.kdv_dahil_tutar),
            g.belge_no or "",
        ])

    _auto_width(ws)
    resp = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    resp["Content-Disposition"] = "attachment; filename=gider_raporu.xlsx"
    wb.save(resp)
    return resp


# ── KDV Beyanname Özeti ───────────────────────────────────────────────────
def excel_kdv_beyanname(request):
    """
    Fatura kalemlerini KDV oranına göre gruplandırarak
    KDV-1 beyanname ozeti oluşturur.
    Parametreler: ?yil=2025&ay=1  (ay belirtilmezse tüm yıl)
    """
    from collections import defaultdict
    from django.db.models import Sum

    yil = request.GET.get("yil") or str(timezone.now().year)
    ay = request.GET.get("ay") or ""

    try:
        yil_int = int(yil)
    except ValueError:
        yil_int = timezone.now().year

    filtre = {"fatura__durum__in": ["kesildi", "odendi"]}
    filtre["fatura__tarih__year"] = yil_int
    if ay:
        try:
            filtre["fatura__tarih__month"] = int(ay)
        except ValueError:
            pass

    # Satış faturalarından KDV özetleri
    satis_qs = (
        FaturaKalemi.objects
        .filter(**filtre, fatura__tip="satis")
        .values("kdv_orani")
        .annotate(
            matrah=Sum("kdv_haric_tutar"),
            kdv=Sum("kdv_tutari"),
            toplam=Sum("kdv_dahil_tutar"),
        )
        .order_by("kdv_orani")
    )

    # Alış faturalarından KDV özetleri (indirilecek KDV)
    alis_qs = (
        FaturaKalemi.objects
        .filter(**filtre, fatura__tip="alis")
        .values("kdv_orani")
        .annotate(
            matrah=Sum("kdv_haric_tutar"),
            kdv=Sum("kdv_tutari"),
            toplam=Sum("kdv_dahil_tutar"),
        )
        .order_by("kdv_orani")
    )

    wb = openpyxl.Workbook()

    # ── Sayfa 1: Satış KDV ─────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Hesaplanan KDV (Satış)"

    donem = f"{yil_int}/{ay.zfill(2) if ay else 'Tümü'}"
    ws1["A1"] = f"KDV Beyanname Özeti — Hesaplanan KDV (Satış) — Dönem: {donem}"
    ws1["A1"].font = Font(bold=True, size=12)
    ws1.merge_cells("A1:E1")

    _header_row(ws1, ["KDV Oranı (%)", "Matrah (KDV Hariç)", "Hesaplanan KDV", "KDV Dahil Toplam", "Açıklama"], row=3)

    toplam_satis_matrah = Decimal("0")
    toplam_satis_kdv = Decimal("0")
    for row_num, row in enumerate(satis_qs, 4):
        matrah = row["matrah"] or Decimal("0")
        kdv = row["kdv"] or Decimal("0")
        toplam = row["toplam"] or Decimal("0")
        toplam_satis_matrah += matrah
        toplam_satis_kdv += kdv
        aciklama = f"%{row['kdv_orani']} KDV'li satışlar"
        ws1.append([f"%{row['kdv_orani']}", float(matrah), float(kdv), float(toplam), aciklama])

    # Toplam satırı
    fill_toplam = PatternFill(fill_type="solid", fgColor="FFF3CD")
    toplam_row_idx = ws1.max_row + 1
    ws1.cell(toplam_row_idx, 1, "TOPLAM").font = Font(bold=True)
    ws1.cell(toplam_row_idx, 2, float(toplam_satis_matrah)).font = Font(bold=True)
    ws1.cell(toplam_row_idx, 3, float(toplam_satis_kdv)).font = Font(bold=True)
    for c in ws1[toplam_row_idx]:
        c.fill = fill_toplam

    ws1.append([])
    ws1.append(["Not: Fatura durumu 'Kesildi' veya 'Ödendi' olanlar dahil edilmiştir."])

    _auto_width(ws1)

    # ── Sayfa 2: Alış KDV ─────────────────────────────────────────────────
    ws2 = wb.create_sheet("İndirilecek KDV (Alış)")

    ws2["A1"] = f"KDV Beyanname Özeti — İndirilecek KDV (Alış) — Dönem: {donem}"
    ws2["A1"].font = Font(bold=True, size=12)
    ws2.merge_cells("A1:E1")

    _header_row(ws2, ["KDV Oranı (%)", "Matrah (KDV Hariç)", "İndirilecek KDV", "KDV Dahil Toplam", "Açıklama"], row=3)

    toplam_alis_matrah = Decimal("0")
    toplam_alis_kdv = Decimal("0")
    for row_num, row in enumerate(alis_qs, 4):
        matrah = row["matrah"] or Decimal("0")
        kdv = row["kdv"] or Decimal("0")
        toplam = row["toplam"] or Decimal("0")
        toplam_alis_matrah += matrah
        toplam_alis_kdv += kdv
        aciklama = f"%{row['kdv_orani']} KDV'li alışlar"
        ws2.append([f"%{row['kdv_orani']}", float(matrah), float(kdv), float(toplam), aciklama])

    toplam_row_idx2 = ws2.max_row + 1
    ws2.cell(toplam_row_idx2, 1, "TOPLAM").font = Font(bold=True)
    ws2.cell(toplam_row_idx2, 2, float(toplam_alis_matrah)).font = Font(bold=True)
    ws2.cell(toplam_row_idx2, 3, float(toplam_alis_kdv)).font = Font(bold=True)
    for c in ws2[toplam_row_idx2]:
        c.fill = fill_toplam

    _auto_width(ws2)

    # ── Sayfa 3: Özet ─────────────────────────────────────────────────────
    ws3 = wb.create_sheet("Beyanname Özeti")
    ws3["A1"] = f"KDV Beyanname Özeti — Dönem: {donem}"
    ws3["A1"].font = Font(bold=True, size=13)
    ws3.merge_cells("A1:B1")

    ws3.append([])
    ws3.append(["Hesaplanan KDV (Satış)", float(toplam_satis_kdv)])
    ws3.append(["İndirilecek KDV (Alış)", float(toplam_alis_kdv)])
    ws3.append([])
    odenmesi_gereken = toplam_satis_kdv - toplam_alis_kdv
    ode_row = ws3.max_row + 1
    ws3.cell(ode_row, 1, "Ödenecek / İade Edilecek KDV").font = Font(bold=True, size=11)
    val_cell = ws3.cell(ode_row, 2, float(odenmesi_gereken))
    val_cell.font = Font(bold=True, size=11, color="DC2626" if odenmesi_gereken > 0 else "16A34A")

    ws3.append([])
    ws3.append([f"Dönem: {donem}", f"Oluşturulma: {timezone.now().strftime('%d.%m.%Y %H:%M')}"])
    ws3.column_dimensions["A"].width = 35
    ws3.column_dimensions["B"].width = 20

    fname = f"kdv_beyanname_{yil_int}{'_ay' + ay if ay else ''}.xlsx"
    resp = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    resp["Content-Disposition"] = f"attachment; filename={fname}"
    wb.save(resp)
    return resp


# ── Dönemsel Özet (HTML görünüm) ──────────────────────────────────────────
def donemsel_ozet(request):
    from fatura.models import Fatura
    from ceksenet.models import CekSenet
    from collections import defaultdict

    yil = int(request.GET.get("yil", timezone.now().year))
    aylar = list(range(1, 13))
    ay_isimleri = [
        "", "Ocak", "Şubat", "Mart", "Nisan", "Mayıs", "Haziran",
        "Temmuz", "Ağustos", "Eylül", "Ekim", "Kasım", "Aralık",
    ]

    data = []
    toplam_gelir = Decimal("0")
    toplam_gider_kdv_haric = Decimal("0")
    toplam_tahsilat = Decimal("0")

    for ay in aylar:
        satis_fat = Fatura.objects.filter(
            tip="satis", durum__in=["kesildi", "odendi"],
            tarih__year=yil, tarih__month=ay,
        )
        satis = sum(f.genel_toplam for f in satis_fat) or Decimal("0")

        alis_fat = Fatura.objects.filter(
            tip="alis", durum__in=["kesildi", "odendi"],
            tarih__year=yil, tarih__month=ay,
        )
        alis = sum(f.genel_toplam for f in alis_fat) or Decimal("0")

        gid = Gider.objects.filter(tarih__year=yil, tarih__month=ay)
        gider_sum = sum(g.kdv_dahil_tutar for g in gid) or Decimal("0")

        tahsilat_sum = sum(
            t.tutar for t in Tahsilat.objects.filter(tarih__year=yil, tarih__month=ay)
        ) or Decimal("0")

        net = satis - alis - gider_sum

        toplam_gelir += satis
        toplam_gider_kdv_haric += gider_sum
        toplam_tahsilat += tahsilat_sum

        data.append({
            "ay_no": ay,
            "ay_adi": ay_isimleri[ay],
            "satis": satis,
            "alis": alis,
            "gider": gider_sum,
            "tahsilat": tahsilat_sum,
            "net": net,
        })

    context = {
        "yil": yil,
        "onceki_yil": yil - 1,
        "sonraki_yil": yil + 1,
        "data": data,
        "toplam_gelir": toplam_gelir,
        "toplam_gider": toplam_gider_kdv_haric,
        "toplam_tahsilat": toplam_tahsilat,
        "net_genel": toplam_gelir - toplam_gider_kdv_haric,
    }
    return render(request, "rapor/donemsel_ozet.html", context)


# ── Rapor Ana Sayfası ─────────────────────────────────────────────────────
def rapor_index(request):
    cariler = Cari.objects.all().order_by("ad")
    return render(request, "rapor/index.html", {
        "cariler": cariler,
        "bugun": timezone.now().date(),
    })
